import re
import json

import openpyxl
from openpyxl.styles import Font

from src.excel import Excel
from src.excel import Sheet
from src.excel import BodyItem


def set_border(worksheet):
    """
    加外框线
    :param worksheet:
    :return:
    """
    border = openpyxl.styles.Side(style='thin', color='000000')
    border_style = openpyxl.styles.Border(left=border, right=border, top=border, bottom=border)
    for row in worksheet.rows:
        for cell in row:
            if cell == worksheet.cell(row=1, column=1):
                continue
            if cell.value or (cell.coordinate in worksheet.merged_cells):
                cell.border = border_style


def create_sheet_desc(worksheet, excel, sheet):
    """
    创建表格sheet 头部信息
    :param worksheet:
    :param excel:
    :param sheet:
    :return:
    """
    pattern_fill = openpyxl.styles.PatternFill(start_color=excel.bg_color, end_color=excel.bg_color, fill_type='solid')
    bg_font = Font(color=excel.has_bg_font_color, name=excel.has_bg_font_name, size=excel.has_bg_font_size, bold=True)

    worksheet.cell(row=1, column=1).value = '=HYPERLINK("#目录!A1", "返回")'
    worksheet.cell(row=1, column=1).style = 'Hyperlink'

    # 写入表头
    name_cell = worksheet.cell(row=3, column=1, value='接口名称')
    name_cell.fill = pattern_fill
    name_cell.font = bg_font
    worksheet.cell(row=3, column=2, value=sheet.name)

    detail_cell = worksheet.cell(row=4, column=1, value='接口说明')
    detail_cell.fill = pattern_fill
    detail_cell.font = bg_font
    worksheet.cell(row=4, column=2, value=sheet.detail)

    url_cell = worksheet.cell(row=5, column=1, value='请求URL')
    url_cell.fill = pattern_fill
    url_cell.font = bg_font
    worksheet.cell(row=5, column=2, value=excel.host + sheet.url)

    method_cell = worksheet.cell(row=6, column=1, value='请求方式')
    method_cell.fill = pattern_fill
    method_cell.font = bg_font
    worksheet.cell(row=6, column=2, value=sheet.method)

    # 合并单元格
    worksheet.merge_cells('B3:G3')
    worksheet.merge_cells('B4:G4')
    worksheet.merge_cells('B5:G5')
    worksheet.merge_cells('B6:G6')


def create_sheet_request(worksheet, excel, req):
    worksheet.cell(row=8, column=1, value='输入参数').font = Font(name=excel.has_bg_font_name)
    worksheet.merge_cells('A8:G8')

    pattern_fill = openpyxl.styles.PatternFill(start_color=excel.bg_color, end_color=excel.bg_color, fill_type='solid')
    bg_font = Font(color=excel.has_bg_font_color, name=excel.has_bg_font_name, size=excel.has_bg_font_size, bold=True)
    border = openpyxl.styles.Side(style='thin', color='000000')
    border_style = openpyxl.styles.Border(left=border, right=border, top=border, bottom=border)

    for i, header in enumerate(['标签名称', '中文说明', '字段类型', '字典长度', '是否必填', '验证方式', '备注'],
                               start=1):
        worksheet.cell(row=9, column=i, value=header).font, \
            worksheet.cell(row=9, column=i, value=header).fill = bg_font, pattern_fill
    req_data = create_child_cell(req)
    for i in range(len(req_data)):
        worksheet.cell(row=10 + i, column=1, value=req_data[i].deep * "  " + req_data[i].name)
        worksheet.cell(row=10 + i, column=2, value=req_data[i].desc)
        worksheet.cell(row=10 + i, column=3, value=req_data[i].type)
        worksheet.cell(row=10 + i, column=5, value='是' if req_data[i].require else '否')
        worksheet.cell(row=10 + i, column=7, value=req_data[i].desc)
        for j in range(1, 8):
            worksheet.cell(row=10 + i, column=j).border = border_style


def create_sheet_response(worksheet, excel, resp):
    worksheet.cell(row=8, column=9, value='返回参数').font = Font(name=excel.has_bg_font_name)
    worksheet.merge_cells('I8:Q8')

    pattern_fill = openpyxl.styles.PatternFill(start_color=excel.bg_color, end_color=excel.bg_color, fill_type='solid')
    bg_font = Font(color=excel.has_bg_font_color, name=excel.has_bg_font_name, size=excel.has_bg_font_size, bold=True)
    border = openpyxl.styles.Side(style='thin', color='000000')
    border_style = openpyxl.styles.Border(left=border, right=border, top=border, bottom=border)

    for i, header in enumerate(['序号', '标签名称', '中文说明', '字段类型', '字典长度', '循环', '是否必填',
                                '使用权限（无说明为管理员用户共用）', '备注'], start=9):
        worksheet.cell(row=9, column=i, value=header).font, \
            worksheet.cell(row=9, column=i, value=header).fill = bg_font, pattern_fill

    req_data = create_child_cell(resp)
    for i in range(len(req_data)):
        resp_map = excel.resp_map.get(req_data[i].name, {})
        worksheet.cell(row=10 + i, column=9, value=i + 1)
        worksheet.cell(row=10 + i, column=10, value=req_data[i].deep * "  " + req_data[i].name)
        worksheet.cell(row=10 + i, column=11, value=resp_map.get('chinese', req_data[i].desc))
        worksheet.cell(row=10 + i, column=12, value=req_data[i].type)
        worksheet.cell(row=10 + i, column=15, value='是' if resp_map.get('require', req_data[i].require) else '否')
        worksheet.cell(row=10 + i, column=17, value=resp_map.get('desc', req_data[i].desc))
        for j in range(9, 18):
            worksheet.cell(row=10 + i, column=j).border = border_style


def create_child_cell(datas, deep=0):
    if datas is None:
        return []
    d = []
    for data in datas:
        children = create_child_cell(data.get('children'), deep + 1)
        desc = data.get('description')
        require = data.get('require')
        name = data.get('name')
        type = data.get('type')
        d.append(BodyItem(children, desc, require, name, type, deep))
        d.extend(children)
    return d


def create_sheet(workbook, excel: Excel, sheet: Sheet):
    """
    创建
    :param workbook:
    :param excel:
    :param sheet:
    :return:
    """

    worksheet = workbook.create_sheet(title=sheet.title)

    # 创建头部信息
    create_sheet_desc(worksheet, excel, sheet)

    # 创建请求信息
    create_sheet_request(worksheet, excel, sheet.req)
    # 创建返回值信息
    create_sheet_response(worksheet, excel, sheet.resp)

    # 给表格加上外框线
    set_border(worksheet)
    # 调整A列宽度以适应最宽的单元格内容
    for col in worksheet.columns:
        worksheet.column_dimensions[col[0].column_letter].auto_fit = True


def get_data(file_name):
    """
    解析html文件获取接口数据
    :param file_name:
    :return:
    """
    with open(file_name, 'r', encoding='utf-8') as f:
        findall = re.findall("var datas=(.*);\n", f.read())
        if len(findall) != 1:
            exit(0)
        return json.loads(findall[0])


def create_menu(workbook, excel, toc_list: list):
    """
    创建目录
    :param workbook:
    :param excel:
    :param toc_list:
    :return:
    """
    workbook.active.title = "目录"
    worksheet = workbook.active

    pattern_fill = openpyxl.styles.PatternFill(start_color=excel.bg_color, end_color=excel.bg_color, fill_type='solid')
    bg_font = Font(color=excel.has_bg_font_color, name=excel.has_bg_font_name, size=excel.has_bg_font_size, bold=True)
    border = openpyxl.styles.Side(style='thin', color='000000')
    border_style = openpyxl.styles.Border(left=border, right=border, top=border, bottom=border)

    for i, header in enumerate(['编号', '产品', '页面名', '事件', '接口'], start=1):
        worksheet.cell(row=2, column=i, value=header).font, worksheet.cell(row=2, column=i,
                                                                           value=header).fill = bg_font, pattern_fill
    for i, j in enumerate(toc_list, start=3):
        worksheet.cell(row=i, column=1, value=i - 2)
        worksheet.cell(row=i, column=2, value=excel.file_name)
        worksheet.cell(row=i, column=3, value=j.get('pt'))
        worksheet.cell(row=i, column=4, value=j.get('t'))
        cell = worksheet.cell(row=i, column=5, value=j.get('t'))
        cell.value = f"=HYPERLINK(\"#\'{j.get('t')}\'!A1\",\"{j.get('t')}\")"
        cell.style = 'Hyperlink'
        # 给表格加上外框线

    set_border(worksheet)
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 35
    worksheet.column_dimensions['D'].width = 40
    worksheet.column_dimensions['E'].width = 40


def create_excel(data, title, host):
    """
    创建一个新的Excel文件
    :param data:
    :return:
    """
    workbook = openpyxl.Workbook()

    instance = data.get('instance')
    title = title if title else instance.get('title')
    file_desc = instance.get('description')

    tags = data.get('tags')
    excel = Excel(title, host)

    toc_list = []
    for sheet in get_sheets(tags):
        # 轮流创建页
        create_sheet(workbook, excel, sheet)
        toc_list.append({'t': sheet.title, 'pt': sheet.parent_title})

    # 创建目录
    create_menu(workbook, excel, toc_list)

    # 保存Excel文件
    workbook.save(excel.file_name + '.xlsx')


def get_sheets(tags):
    """
    获取页
    :param tags:
    :return:
    """
    for tag in tags:
        one_title = tag.get('name')
        one_description = tag.get('description')
        for child in tag.get('childrens'):
            desc = child.get('description').replace("<p>", "").replace("</p>", "")
            method = child.get('methodType')
            url = child.get('showUrl')
            summary = child.get('summary').replace('[', '【').replace(']', '】')
            produces = child.get('produces')
            req_params = child.get('reqParameters')
            resp_params = child.get('multipData').get('responseParameters')
            yield Sheet(parent_title=one_title, title=summary, name=summary, detail=desc, url=url, method=method,
                        req=req_params,
                        resp=resp_params)


if __name__ == '__main__':
    # title = '分析研判'
    # host = 'http://ip:port'
    title = input("输入文档名称(默认按照swagger内的名称)：")
    host = input("输入host(默认http://ip:port)：") or "http://ip:port"
    create_excel(get_data('swagger.html'), title=title, host=host)



